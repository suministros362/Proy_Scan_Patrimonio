from datetime import datetime
import openpyxl
from openpyxl.styles import Alignment
import os

import xlrd
from xlutils.copy import copy

class PlanillaGenerator:
    def __init__(self, ruta_plantilla_relevo=os.path.join("plantillas", "plantilla_relevo.xlsx"), ruta_plantilla_busqueda=os.path.join("plantillas", "plantilla_busqueda.xls")):
        self.ruta_plantilla_relevo = ruta_plantilla_relevo
        self.ruta_plantilla_busqueda = ruta_plantilla_busqueda

    def generar_relevo(self, datos_encabezado: dict, productos: list, ruta_salida: str) -> bool:
        """Llena la plantilla de relevo con los datos del formulario y los productos de la sesión."""
        try:
            wb = openpyxl.load_workbook(self.ruta_plantilla_relevo)
            ws = wb.active

            # 1. Insertar Encabezados y Fecha Actual
            # 📌 Cambia las celdas ("B2", "B3", etc.) por las reales de tu Excel
            ws["H5"] = datetime.now().strftime("%d/%m/%Y")
            ws["C8"] = datos_encabezado.get("piso", "")
            ws["C9"] = datos_encabezado.get("oficina", "")
            ws["C10"] = datos_encabezado.get("area", "")
            ws["C11"] = datos_encabezado.get("dependencia", "")
            ws["F8"] = datos_encabezado.get("responsable", "")
            ws["F9"] = datos_encabezado.get("subresponsable", "")

            # Configuración de alineación con ajuste de texto
            alineacion_centrado = Alignment(horizontal="center", vertical="center", wrap_text=True)
            alineacion_texto = Alignment(horizontal="left", vertical="center", wrap_text=True)

            ws["H5"].alignment = alineacion_centrado
            ws["C8"].alignment = alineacion_centrado
            ws["C9"].alignment = alineacion_centrado    
            ws["C10"].alignment = alineacion_centrado
            ws["C11"].alignment = alineacion_centrado

            # 2. Insertar Listado de Bienes
            fila_inicio = 14  # 📌 Ajusta la fila inicial de tu tabla

            for idx, prod in enumerate(productos, start=1):
                fila = fila_inicio + idx - 1

                ws[f"A{fila}"] = idx
                ws[f"B{fila}"] = prod.nro_inventario
                ws[f"C{fila}"] = prod.nro_nuevo
                ws[f"D{fila}"] = prod.elemento
                ws[f"E{fila}"] = prod.marca
                ws[f"F{fila}"] = prod.modelo
                ws[f"G{fila}"] = prod.nro_serie
                ws[f"H{fila}"] = prod.observaciones  # Agregando la columna de observaciones
                ws[f"I{fila}"] = prod.sector.upper()  # Agregando la columna de sector

                # Aplicar alineación
                ws[f"A{fila}"].alignment = alineacion_centrado
                ws[f"B{fila}"].alignment = alineacion_centrado
                ws[f"C{fila}"].alignment = alineacion_centrado
                ws[f"D{fila}"].alignment = alineacion_texto
                ws[f"E{fila}"].alignment = alineacion_texto
                ws[f"F{fila}"].alignment = alineacion_texto
                ws[f"G{fila}"].alignment = alineacion_centrado
                ws[f"H{fila}"].alignment = alineacion_texto  # Alineación para observaciones
                ws[f"I{fila}"].alignment = alineacion_texto  # Alineación para sector

            # 3. Guardar en el destino elegido
            wb.save(ruta_salida)
            return True

        except Exception as e:
            print(f"Error al generar la planilla de relevo: {e}")
            return False

    def generar_busqueda(self, resultados: list, ruta_salida: str) -> bool:
        """
        Genera la planilla en formato .xls (Excel 97-2003).
        Escribe los códigos en Columna A y luego los repite en Columna B 
        a continuación de la última fila cargada en A.
        """
        try:
            if not os.path.exists(self.ruta_plantilla_busqueda):
                print(f"Error: No se encontró la plantilla .xls en {self.ruta_plantilla_busqueda}")
                return False

            # 1. Cargar el libro .xls existente (formatting_info=True conserva fuentes y estilos)
            rb = xlrd.open_workbook(self.ruta_plantilla_busqueda, formatting_info=True)
            
            # 2. Crear una copia modificable usando xlutils
            wb = copy(rb)
            ws = wb.get_sheet(0)  # Primera hoja

            # -------------------------------------------------------------
            # 📌 NOTA: xlrd/xlwt usan índices base cero (0 = Columna A / Fila 1 = Fila 0)
            # -------------------------------------------------------------

            # 4. Volcado de datos (Ejemplo: Fila 8 en Excel es índice 7)
            fila_inicio_A = 3  # ⚠️ Fila 8 de Excel (base 0)
            cantidad_elementos = len(resultados)

            # A) Pegar códigos en Columna A (Columna 0)
            for i, item in enumerate(resultados):
                fila_actual_A = fila_inicio_A + i
                codigo = getattr(item, "nro_nuevo", "")
                ws.write(fila_actual_A, 0, codigo)  # Columna 0 = Columna A

            # B) Pegar los MISMOS códigos en Columna B (Columna 1), a continuación de A
            fila_inicio_B = fila_inicio_A + cantidad_elementos

            for i, item in enumerate(resultados):
                fila_actual_B = fila_inicio_B + i
                codigo = getattr(item, "nro_nuevo", "")
                ws.write(fila_actual_B, 1, codigo)  # Columna 1 = Columna B

            # 5. Guardar manteniendo el formato .xls
            # Asegurarse de que la extensión sea .xls
            if not ruta_salida.lower().endswith(".xls"):
                ruta_salida = os.path.splitext(ruta_salida)[0] + ".xls"

            wb.save(ruta_salida)
            return True

        except Exception as e:
            print(f"Error al generar la planilla .xls: {e}")
            return False